import json
import pickle
import regex as re
from lxml import etree
from pattern.nl import parse, conjugate
import openpyxl
import os
import string
import spacy
ennlp = spacy.load("en_core_web_lg")
nlnlp = spacy.load("nl_core_news_lg")
import copy

currentpath = os.getcwd()

with open(currentpath + '/Data/Relevantinfo.pkl', 'rb') as f:
    relevantinfo = pickle.load(f)

with open(currentpath + '/Data/FilenamelistNewSplit.pkl', 'rb') as f:
    filenames = pickle.load(f)

with open(currentpath + '/Data/AlldatasetsNewSplit2.pkl', 'rb') as f:
    alldatasets = pickle.load(f)

def gettitle(text):
    for paragraph in text:
        if len(paragraph) == 0:
            continue
        for sentence in paragraph:
            if 'title' in sentence:
                title = sentence['title']
                if title == '':
                    dataset = sentence['sentenceidx'][0]
                    text = sentence['sentenceidx'][1]
                    filename = filenames[dataset][text]
                    title = filename
                return title

def squaddataset(alldatasets):
    datasetlist = ['Train', 'Dev', 'Test']
    for datasetidx, dataset in enumerate(alldatasets):
        alldict = {'version': 1.0, 'data': []}
        for text in dataset:
            title = gettitle(text)
            textdict = {'title': title, 'paragraphs': []}
            for paragraph in text:
                if len(paragraph) == 0:
                    continue
                paragraphtext = [x['sentence'] for x in paragraph]
                paragraphtext = ' '.join(paragraphtext)
                paragraphdict = {'qas': [], 'context': paragraphtext}
                questiondict = {}
                for sentence in paragraph:
                    currentsentence = sentence['sentence'].strip()

                    #The paragraphidx key and textidx keys are only present in the Prodigy annotated sentences; MyTool has a list for sentenceidx
                    if 'paragraphidx' in sentence:
                        annotationtype = 'Prodigy'
                    else:
                        annotationtype = 'MyTool'
                        startidx = re.search(re.escape(currentsentence), paragraphtext).start()

                    for combineddictkey in sentence:
                        if (combineddictkey != 'sentenceidx') and (combineddictkey != 'filename') and (combineddictkey != 'sentence') and (combineddictkey != 'title') and (combineddictkey != 'textidx') and (combineddictkey != 'paragraphidx') and (combineddictkey != 'datasetidx'):
                            answerdict = []

                            if annotationtype == 'Prodigy':
                                for occurrence in sentence[combineddictkey]:
                                    ad = {'text': occurrence[0], 'answer_start': occurrence[3]}
                                    answerdict.append(ad)

                            elif annotationtype == 'MyTool':
                                if ';;' in sentence[combineddictkey]:
                                    newsplit = sentence[combineddictkey].split(';;')
                                    newsplit = [x.strip() for x in newsplit]
                                    for content in newsplit:
                                        startindices = []
                                        sentencestartidx = re.finditer(r'\b' + re.escape(content) + r'\b', currentsentence)
                                        for sentstartidx in sentencestartidx:
                                            actualidx = startidx + sentstartidx.start()
                                            startindices.append(actualidx)
                                        #sentencestartidx = re.search(r'\b' + re.escape(content) + r'\b', currentsentence).start()
                                        #actualidx = startidx + sentencestartidx
                                        for actualidx in startindices:
                                            ad = {'text': content, 'answer_start': actualidx}
                                            if (combineddictkey in questiondict) and (ad in questiondict[combineddictkey]['answers']):
                                                continue
                                            else:
                                                answerdict.append(ad)
                                else:
                                    startindices = []
                                    sentencestartidx = re.finditer(r'\b' + re.escape(sentence[combineddictkey]) + r'\b', currentsentence)
                                    for sentstartidx in sentencestartidx:
                                        actualidx = startidx + sentstartidx.start()
                                        startindices.append(actualidx)

                                    if len(startindices) > 0:
                                        for actualidx in startindices:
                                            ad = {'text': sentence[combineddictkey], 'answer_start': actualidx}
                                            if (combineddictkey in questiondict) and (ad in questiondict[combineddictkey]['answers']):
                                                continue
                                            else:
                                                answerdict.append(ad)
                                        #sentencestartidx = re.search(r'\b' + re.escape(sentence[combineddictkey]) + r'\b', currentsentence).start()
                                        #actualidx = startidx + sentencestartidx
                                        #answerdict = [{'text': sentence[combineddictkey], 'answer_start': actualidx}]

                                    else:
                                        actualidx = startidx
                                        ad = {'text': currentsentence, 'answer_start': actualidx}
                                        if (combineddictkey in questiondict) and (answerdict in questiondict[combineddictkey]['answers']):
                                            pass
                                        else:
                                            answerdict.append(ad)


                            if combineddictkey not in questiondict:
                                if annotationtype == 'MyTool':
                                    questionid = str(sentence['sentenceidx'][0]) + '/' + str(sentence['sentenceidx'][1]) + '/' + str(sentence['sentenceidx'][2]) + '_' + combineddictkey
                                elif annotationtype == 'Prodigy':
                                    questionid = str(sentence['textidx']) + '/' + str(sentence['paragraphidx']) + '/' + str(sentence['sentenceidx']) + '_' + combineddictkey

                                questiondict.update({combineddictkey: {'question': combineddictkey, 'id': questionid, 'answers': answerdict, 'is_impossible': False}})
                            else:
                                questiondict[combineddictkey]['answers'] = questiondict[combineddictkey]['answers'] + answerdict

                for information in relevantinfo:
                    if information not in questiondict:
                        if annotationtype == 'MyTool':
                            questionid = str(paragraph[0]['sentenceidx'][0]) + '/' + str(paragraph[0]['sentenceidx'][1]) + '/' + str(paragraph[0]['sentenceidx'][2]) + '_' + information
                        elif annotationtype == 'Prodigy':
                            questionid = str(paragraph[0]['textidx']) + '/' + str(paragraph[0]['paragraphidx']) + '/' + str(paragraph[0]['sentenceidx']) + '_' + information

                        questiondict.update(
                            {information: {'question': information, 'id': questionid, 'answers': [], 'is_impossible': True}})

                for question in questiondict:
                    paragraphdict['qas'].append(questiondict[question])
                textdict['paragraphs'].append(paragraphdict)
            alldict['data'].append(textdict)

        currentpath = os.getcwd()
        with open(currentpath + '/Data/SQUADFormat' + datasetlist[datasetidx] + '.json', 'w') as outfile:
            json.dump(alldict, outfile)

def find_sub_list(sl,l):
    sll=len(sl)
    for ind in (i for i,e in enumerate(l) if e==sl[0]):
        if l[ind:ind+sll]==sl:
            return ind,ind+sll-1

def find_sub_list_multiple(sl,l):
    results=[]
    sll=len(sl)
    for ind in (i for i,e in enumerate(l) if e==sl[0]):
        if l[ind:ind+sll]==sl:
            results.append((ind,ind+sll-1))

    return results

def sentenceindices(splitsentence, referencedict):
    sentencereplaceindexdict = {}
    for reference in referencedict:
        referenceword = re.search(r'^(.*?)__', reference).group(1)
        if referenceword != 'True':
            try:
                referenceidx = re.search(r'__(\d+)$', reference).group(1)
            except AttributeError:
                print(reference)
                print(splitsentence)
                exit(1)

            parsesentence = parse(referenceword, tagset='WOTAN')
            parsesentence = re.split(r'\s|\n', parsesentence)
            splitsubsentence = []
            for word in parsesentence:
                wordsplit = re.search(r'^(.*?)/(.*?)$', word)
                splitsubsentence.append(wordsplit.group(1))

            sublistindex = find_sub_list_multiple(splitsubsentence, splitsentence)
            try:
                subdata = sublistindex[int(referenceidx)]
                newsublistindex = list(range(subdata[0], subdata[1]+1))
            except (TypeError, IndexError) as e:
                try:
                    subdata = sublistindex[0]
                    newsublistindex = list(range(subdata[0], subdata[1] + 1))
                except (TypeError, IndexError) as e:
                    print(splitsubsentence)
                    print(splitsentence)
                    print(reference)
                    exit(2)

            sentencereplaceindexdict.update({referencedict[reference][2]: newsublistindex})

    return sentencereplaceindexdict

def lexicalizesentence(sentence, referencedict, annotationtype):
    #HIER NOG EEN PRODIGYVARIANT VAN MAKEN!
    if annotationtype == 'Prodigy':
        print('MAKE PRODIGY VARIANT')
        exit(0)

    #lowersentence = sentence.lower()
    if annotationtype == 'MyTool':
        parsesentence = parse(sentence, tagset='WOTAN')

        parsesentence = re.split(r'\s|\n', parsesentence)

        newsentence = []

        splitsentence = []

        for word in parsesentence:
            #Split the word from the morphological information
            wordsplit = re.search(r'^(.*?)/(.*?)$', word)
            try:
                splitsentence.append(wordsplit.group(1))
            except AttributeError:
                print([word])
                print(parsesentence)
                exit(1)
            #Split the wordtype (e.g. verb, noun, determiner) from the morphological information
            wordtype = re.search(r'^(.*?)(\((.*?))/', wordsplit.group(2))
            if wordtype == None:
                newsentence.append(wordsplit.group(1).lower())
            #"Art" are determiners
            elif wordtype.group(1) == 'Art':
                #Remove the brackets, and split the three-part information
                article = re.sub(r'[()]', '', wordtype.group(2))
                article = article.split(',')
                articlestring = 'ART[form=' + article[0] + ',gender=' + article[1] + ',case=' + article[2] + '] ' + wordsplit.group(1).lower()
                newsentence.append(articlestring)
            #"V" is verb
            elif wordtype.group(1) == 'V':
                # Remove the brackets, and split the four-part information
                verb = re.sub(r'[()]', '', wordtype.group(2))
                verb = verb.split(',')
                if len(verb) == 4:
                    verbstring = 'V[type=' + verb[0] + ',tense=' + verb[1] + ',person=' + verb[2] + ',number=' + verb[3] + '] ' + conjugate(wordsplit.group(1), 'INFINITIVE').lower()
                elif len(verb) == 3:
                    verbstring = 'V[type=' + verb[0] + ',tense=' + verb[1] + ',person=' + verb[2] + '] ' + conjugate(wordsplit.group(1), 'INFINITIVE').lower()
                elif len(verb) == 2:
                    verbstring = 'V[type=' + verb[0] + ',tense=' + verb[1] + '] ' + conjugate(wordsplit.group(1), 'INFINITIVE').lower()
                newsentence.append(verbstring)
            else:
                newsentence.append(wordsplit.group(1).lower())
        # print(wordtype.group(1), wordtype.group(2))

    #newsentence = ' '.join(newsentence)
    sentencereplaceindexdict = sentenceindices(splitsentence, referencedict)

    for reference in sentencereplaceindexdict:
        for idx, sentenceidx in enumerate(sentencereplaceindexdict[reference]):
            if idx == 0:
                newsentence[sentenceidx] = reference
                splitsentence[sentenceidx] = reference
            else:
                newsentence[sentenceidx] = ''
                splitsentence[sentenceidx] = ''

    newsentence = [x for x in newsentence if x != '']
    newsentence = ' '.join(newsentence)

    splitsentence = [x for x in splitsentence if x != '']
    splitsentence = ' '.join(splitsentence)
    return newsentence, splitsentence

def templatizesentences(sentence, referencedict, annotationtype):
    if annotationtype == 'MyTool':
        endoc = ennlp(sentence)
        splitsentence = [token.text for token in endoc]
        #newsentence = ' '.join(newsentence)
        sentencereplaceindexdict = sentenceindices(splitsentence, referencedict)

        for reference in sentencereplaceindexdict:
            for idx, sentenceidx in enumerate(sentencereplaceindexdict[reference]):
                if idx == 0:
                    splitsentence[sentenceidx] = reference
                else:
                    splitsentence[sentenceidx] = ''

        splitsentence = [x for x in splitsentence if x != '']
        splitsentence = ' '.join(splitsentence)

        return splitsentence

    elif annotationtype == 'Prodigy':
        sortedreferencedict = sorted(referencedict.items(), key=lambda x: x[1][4], reverse=True)
        newsentence = sentence
        for reference in sortedreferencedict:
            newsentence = newsentence[:reference[1][4]] + reference[1][2] + newsentence[reference[1][5]:]

        return newsentence

def webnlg(alldatasets, language):
    datasetlist = ['Train', 'Dev', 'Test']
    for datasetidx, dataset in enumerate(alldatasets):
        #Start with the root and its child node
        benchmark = etree.Element("benchmark")
        entries = etree.SubElement(benchmark, "entries")
        entityid = 1
        #The datasets contain texts
        for text in dataset:
            #Which contain paragraphs
            for paragraph in text:
                if len(paragraph) == 0:
                    continue
                #Which contain sentences
                for sentence in paragraph:
                    infolength = 0
                    refnum = 1
                    triplesetlist = []
                    sortedtriplesetlist = []
                    referencedict = {}
                    referencelist = []

                    #Paragraphidx is an indicator that it was annotated with prodigy, otherwise it's annotated with my tool
                    if 'paragraphidx' in sentence:
                        annotationtype = 'Prodigy'
                    else:
                        annotationtype = 'MyTool'

                    #Go over each annotated piece of information
                    for combineddictkey in sentence:
                        #The annotation types below don't contain content from the sentence, but are meta information
                        if (combineddictkey != 'sentenceidx') and (combineddictkey != 'filename') and (combineddictkey != 'sentence') and (combineddictkey != 'title') and (combineddictkey != 'textidx') and (combineddictkey != 'paragraphidx') and (combineddictkey != 'datasetidx'):
                            infolength += 1

                            if annotationtype == 'Prodigy':
                                for occurrence in sentence[combineddictkey]:
                                    #Let's take a look if there are determiners before the selected content
                                    if language == 'en':
                                        #If the content doesn't start right at the start
                                        if occurrence[5] != 0:
                                            #Check if the word before the occurrence is a determiner and if that's the case, include it in the annotated string
                                            if (sentence['sentence'][occurrence[5]-4:occurrence[5]] == 'the ') or (sentence['sentence'][occurrence[5]-4:occurrence[5]] == 'The '):
                                                sentencestartidx = occurrence[5]-4
                                            elif (sentence['sentence'][occurrence[5]-3:occurrence[5]] == 'an ') or (sentence['sentence'][occurrence[5]-3:occurrence[5]] == 'An '):
                                                sentencestartidx = occurrence[5]-3
                                            elif (sentence['sentence'][occurrence[5]-2:occurrence[5]] == 'a ') or (sentence['sentence'][occurrence[5]-2:occurrence[5]] == 'A '):
                                                sentencestartidx = occurrence[5]-2
                                            else:
                                                sentencestartidx = occurrence[5]
                                        else:
                                            sentencestartidx = occurrence[5]
                                    elif language == 'nl':
                                        #If the content doesn't start right at the start
                                        if occurrence[5] != 0:
                                            #Check if the word before the occurrence is a determiner
                                            if (sentence['sentence'][occurrence[5]-4:occurrence[5]] == 'het ') or (sentence['sentence'][occurrence[5]-4:occurrence[5]] == 'Het ') or (sentence['sentence'][occurrence[5]-4:occurrence[5]] == 'een ') or (sentence['sentence'][occurrence[5]-4:occurrence[5]] == 'Een '):
                                                sentencestartidx = occurrence[5]-4
                                            elif (sentence['sentence'][occurrence[5]-3:occurrence[5]] == 'de ') or (sentence['sentence'][occurrence[5]-3:occurrence[5]] == 'De '):
                                                sentencestartidx = occurrence[5]-3
                                            else:
                                                sentencestartidx = occurrence[5]
                                        else:
                                            sentencestartidx = occurrence[5]


                                    newattribute = sentence['sentence'][sentencestartidx:occurrence[6]] + '__' + combineddictkey + '__' + str(sentencestartidx)
                                    entity = re.sub(r' ', '_', occurrence[0])

                                    referencedict.update({newattribute: [entity, refnum, 'PATIENT-' + str(refnum), 'description', sentencestartidx, occurrence[6]]})
                                    refnum += 1

                                    newsentence = re.sub(r' ', '_', occurrence[0])
                                    combineddictstring = combineddictkey + ' | ' + newsentence
                                    triplesetlist.append(combineddictstring)
                                    sortedtriplesetlist.append([sentencestartidx, combineddictstring])

                            elif annotationtype == 'MyTool':
                                # ';;' is the method of adding multiple phrases to the same category in a sentence in MyTool. So first, if we want to look at every single instance, we have to split this.
                                if ';;' in sentence[combineddictkey]:
                                    newsplit = sentence[combineddictkey].split(';;')
                                    newsplit = [x.strip() for x in newsplit]

                                    for content in newsplit:
                                        startindices = []
                                        #Search for the start index of all occurrences of the annotated string (and a determiner before if there is one)
                                        if language == 'nl':
                                            sentencestartidx = re.finditer(r'(\s|\b|[' + re.escape(string.punctuation) + r']|$)(((de )|(het )|(een )|(De )|(Het )|(Een ))?' + re.escape(content) + r')(\s|\b|[' + re.escape(string.punctuation) + r']|$)', sentence['sentence'])
                                        elif language == 'en':
                                            sentencestartidx = re.finditer(r'(\s|\b|[' + re.escape(string.punctuation) + r']|$)(((the )|(an )|(a )|(The )|(An )|(A ))?' + re.escape(content) + r')(\s|\b|[' + re.escape(string.punctuation) + r']|$)', sentence['sentence'])
                                        for sentstartidx in sentencestartidx:
                                            startindices.append(sentstartidx.start())
                                        #If no startindices were found, we're making 0 the start index
                                        if len(startindices) == 0:
                                            startidx = 0

                                        #Now we will search for all occurrences of the annotated string (plus optional determiner before) again; this will result in a list of tuples with the full string + determiner in position 1 of the tuple
                                        if language == 'nl':
                                            withattribute = re.findall(r'(\s|\b|[' + re.escape(string.punctuation) + r']|$)(((de )|(het )|(een )|(De )|(Het )|(Een ))?' + re.escape(content) + r')(\s|\b|[' + re.escape(string.punctuation) + r']|$)', sentence['sentence'])
                                        elif language == 'en':
                                            withattribute = re.findall(r'(\s|\b|[' + re.escape(string.punctuation) + r']|$)(((the )|(an )|(a )|(The )|(An )|(A ))?' + re.escape(content) + r')(\s|\b|[' + re.escape(string.punctuation) + r']|$)', sentence['sentence'])
                                        #If no occurrence was found, we will just make a string with the annotated string + '__' + the annotation type
                                        if len(withattribute) == 0:
                                            withattribute = None
                                            newattribute = content + '__' + combineddictkey
                                        else:
                                            #Otherwise we will make a list containing all occurrences of the string + annotation type + the occurrence number to make sure that each value is unique when making a referencedict
                                            newattribute = [x[1] + '__' + combineddictkey + '__' + str(newkeyidx) for newkeyidx, x in enumerate(withattribute)]

                                        #Transform the spaces in the entity to underscores to fit the format of WebNLG
                                        entity = re.sub(r' ', '_', content)

                                        if withattribute == None:
                                            if sentence[combineddictkey] not in referencelist:
                                                referencedict.update({newattribute: [entity, refnum, 'PATIENT-' + str(refnum), 'description']})
                                                referencelist.append(sentence[combineddictkey])
                                                refnum += 1

                                                newsentence = re.sub(r' ', '_', content)
                                                combineddictstring = combineddictkey + ' | ' + newsentence
                                                triplesetlist.append(combineddictstring)
                                                sortedtriplesetlist.append([startidx, combineddictstring])
                                        else:
                                            for naidx, naval in enumerate(newattribute):
                                                if withattribute[naidx][0] + '_' + str(naidx) not in referencelist:
                                                    referencedict.update({naval: [entity, refnum, 'PATIENT-' + str(refnum), 'description']})
                                                    referencelist.append(withattribute[naidx][0] + '_' + str(naidx))
                                                    refnum += 1

                                                    newsentence = re.sub(r' ', '_', content)
                                                    combineddictstring = combineddictkey + ' | ' + newsentence
                                                    triplesetlist.append(combineddictstring)
                                                    sortedtriplesetlist.append([startindices[naidx], combineddictstring])
                                                    break
                                else:
                                    combineddictkey = re.sub('\d+', '', combineddictkey)
                                    #newkey = combineddictkey

                                    startindices = []
                                    if language == 'nl':
                                        sentencestartidx = re.finditer(r'(\s|\b|[' + re.escape(string.punctuation) + r']|$)(((de )|(het )|(een )|(De )|(Het )|(Een ))?' + re.escape(sentence[combineddictkey]) + r')(\s|\b|[' + re.escape(string.punctuation) + r']|$)', sentence['sentence'])
                                    elif language == 'en':
                                        sentencestartidx = re.finditer(r'(\s|\b|[' + re.escape(string.punctuation) + r']|$)(((the )|(an )|(a )|(The )|(An )|(A ))?' + re.escape(sentence[combineddictkey]) + r')(\s|\b|[' + re.escape(string.punctuation) + r']|$)', sentence['sentence'])
                                    for sentstartidx in sentencestartidx:
                                        startindices.append(sentstartidx.start())
                                    if len(startindices) == 0:
                                        startidx = 0

                                    if language == 'nl':
                                        withattribute = re.findall(r'(\s|\b|[' + re.escape(string.punctuation) + r']|$)(((de )|(het )|(een )|(De )|(Het )|(Een ))?' + re.escape(sentence[combineddictkey]) + r')(\s|\b|[' + re.escape(string.punctuation) + r']|$)', sentence['sentence'])
                                    elif language == 'en':
                                        withattribute = re.findall(r'(\s|\b|[' + re.escape(string.punctuation) + r']|$)(((the )|(an )|(a )|(The )|(An )|(A ))?' + re.escape(sentence[combineddictkey]) + r')(\s|\b|[' + re.escape(string.punctuation) + r']|$)', sentence['sentence'])
                                    if len(withattribute) == 0:
                                        withattribute = None
                                        newattribute = sentence[combineddictkey] + '__' + combineddictkey
                                    else:
                                        newattribute = [x[1] + '__' + combineddictkey + '__' + str(newkeyidx) for newkeyidx, x in enumerate(withattribute)]

                                    entity = re.sub(r' ', '_', sentence[combineddictkey])

                                    if withattribute == None:
                                        if sentence[combineddictkey] not in referencelist:
                                            referencedict.update({newattribute: [entity, refnum, 'PATIENT-' + str(refnum), 'description']})
                                            referencelist.append(sentence[combineddictkey])
                                            refnum += 1

                                            newsentence = re.sub(r' ', '_', sentence[combineddictkey])
                                            combineddictstring = combineddictkey + ' | ' + newsentence
                                            triplesetlist.append(combineddictstring)
                                            sortedtriplesetlist.append([startidx, combineddictstring])
                                    else:
                                        for naidx, naval in enumerate(newattribute):
                                            if withattribute[naidx][0] + '_' + str(naidx) not in referencelist:
                                                referencedict.update({naval: [entity, refnum, 'PATIENT-' + str(refnum), 'description']})
                                                referencelist.append(withattribute[naidx][0] + '_' + str(naidx))
                                                refnum += 1

                                                newsentence = re.sub(r' ', '_', sentence[combineddictkey])
                                                combineddictstring = combineddictkey + ' | ' + newsentence
                                                triplesetlist.append(combineddictstring)
                                                try:
                                                    sortedtriplesetlist.append([startindices[naidx], combineddictstring])
                                                except IndexError:
                                                    print(newattribute)
                                                    print(startindices)
                                                    exit(1)
                                                break

                    triplesetlist = sorted(triplesetlist)
                    sortedtriplesetlist = sorted(sortedtriplesetlist, key=lambda x: x[0])
                    sortedtriplesetlist = [x[1] for x in sortedtriplesetlist]

                    if infolength > 0:
                        entry = etree.SubElement(entries, "entry")
                        #PARAMETER CATEGORYNAME TOEVOEGEN
                        entry.set("category", "DutchIncidents")
                        entry.set("eid", "Id" + str(entityid))
                        entityid += 1
                        entry.set("size", str(infolength))
                        originaltripleset = etree.SubElement(entry, "originaltripleset")
                        for triple in triplesetlist:
                            tripleelement = etree.SubElement(originaltripleset, "otriple")
                            tripleelement.text = triple
                        modifiedtripleset = etree.SubElement(entry, "modifiedtripleset")
                        for triple in triplesetlist:
                            modtripleelement = etree.SubElement(modifiedtripleset, "mtriple")
                            modtripleelement.text = triple
                        lex = etree.SubElement(entry, "lex")
                        lex.set("comment", "good")
                        lex.set("lid", "Id1")
                        sortedtripleset = etree.SubElement(lex, "sortedtripleset")
                        sortedtriplesetsentence = etree.SubElement(sortedtripleset, "sentence")
                        sortedtriplesetsentence.set("ID", "1")
                        for triple in sortedtriplesetlist:
                            sortedtripleelement = etree.SubElement(sortedtriplesetsentence, "striple")
                            sortedtripleelement.text = triple
                        references = etree.SubElement(lex, "references")
                        for reference in referencedict:
                            referenceword = re.search(r'^(.*?)__', reference).group(1)
                            if referenceword != 'True':
                                ref = etree.SubElement(references, "reference")
                                ref.set("entity", referencedict[reference][0])
                                ref.set("number", str(referencedict[reference][1]))
                                ref.set("tag", referencedict[reference][2])
                                ref.set("type", referencedict[reference][3])
                                ref.text = referenceword

                        sentencetext = etree.SubElement(lex, "text")
                        sentencetext.text = sentence['sentence']

                        if language == 'nl':
                            lexsentence, templatesentence = lexicalizesentence(sentence['sentence'], referencedict, annotationtype)

                            templatetext = etree.SubElement(lex, "template")
                            templatetext.text = templatesentence
                            lextext = etree.SubElement(lex, "lexicalization")
                            lextext.text = lexsentence

                        elif language == 'en':
                            templatesentence = templatizesentences(sentence['sentence'], referencedict, annotationtype)

                            templatetext = etree.SubElement(lex, "template")
                            templatetext.text = templatesentence

                        entitymap = etree.SubElement(entry, "entitymap")
                        for reference in referencedict:
                            referenceword = re.search(r'^(.*?)__', reference).group(1)
                            if referenceword != 'True':
                                entity = etree.SubElement(entitymap, "entity")
                                entitystring = referencedict[reference][2] + ' | ' + referencedict[reference][0]
                                entity.text = entitystring

        fullxml = etree.tostring(benchmark, xml_declaration=True, pretty_print=True, encoding='utf-8').decode()
        currentpath = os.getcwd()
        with open(currentpath + '/Data/WebNLGFormat' + datasetlist[datasetidx] + '.xml', 'wb') as f:
            f.write(bytes(fullxml, 'UTF-8'))

def wordlistidx_prodigy(occurrence, splitsentence):
    wordidxlist = []

    for word in splitsentence:
        if (word[1] >= occurrence[5]) and (word[1] < occurrence[6]):
            wordidxlist.append(word[2])

    return wordidxlist

def ner(alldatasets, language):
    labels = []

    datasetlist = ['Train', 'Dev', 'Test']

    for datasetidx, dataset in enumerate(alldatasets):
        allsentencelist = []
        allnersentencelist = []
        for text in dataset:
            for paragraph in text:
                if len(paragraph) == 0:
                    continue
                for sentence in paragraph:
                    sublistindexdict = {}
                    if language == 'nl':
                        nldoc = nlnlp(sentence['sentence'])
                        splitsentence = [(token.text, token.idx, token.i) for token in nldoc]
                    elif language == 'en':
                        endoc = ennlp(sentence['sentence'])
                        splitsentence = [(token.text, token.idx, token.i) for token in endoc]
                    nersentence = ['O'] * len(splitsentence)

                    if 'paragraphidx' in sentence:
                        annotationtype = 'Prodigy'
                    else:
                        annotationtype = 'MyTool'

                    for combinedkey in sentence:
                        if (combinedkey != 'sentenceidx') and (combinedkey != 'filename') and (combinedkey != 'sentence') and (combinedkey != 'title') and (combinedkey != 'textidx') and (combinedkey != 'paragraphidx') and (combinedkey != 'datasetidx'):
                            if annotationtype == 'Prodigy':
                                for occurrence in sentence[combinedkey]:
                                    sublistindex = wordlistidx_prodigy(occurrence, splitsentence)
                                    newcombinedkey = combinedkey + '_' + str(occurrence[5])
                                    sublistindexdict.update({newcombinedkey: sublistindex})

                            elif annotationtype == 'MyTool':
                                splitsentence = [x[0] for x in splitsentence]
                                if ';;' in sentence[combinedkey]:
                                    newsplit = sentence[combinedkey].split(';;')
                                    newsplit = [x.strip() for x in newsplit]
                                    for contextidx, content in enumerate(newsplit):
                                        newcombinedkey = combinedkey + '_' + str(contextidx)
                                        if language == 'nl':
                                            nldoc = nlnlp(content)
                                            splitinfo = [token.text for token in nldoc]
                                        elif language == 'en':
                                            endoc = ennlp(content)
                                            splitinfo = [token.text for token in endoc]
                                        sublistindex = find_sub_list_multiple(splitinfo, splitsentence)
                                        for slidx, slval in enumerate(sublistindex):
                                            newsublistindex = list(range(slval[0], slval[1] + 1))
                                            if newsublistindex not in list(sublistindexdict.values()):
                                                newcombinedkey = newcombinedkey + '_' + str(slidx)
                                                sublistindexdict.update({newcombinedkey: newsublistindex})
                                                break
                                else:
                                    if language == 'nl':
                                        nldoc = nlnlp(sentence[combinedkey])
                                        splitinfo = [token.text for token in nldoc]
                                    elif language == 'en':
                                        endoc = ennlp(sentence[combinedkey])
                                        splitinfo = [token.text for token in endoc]
                                    sublistindex = find_sub_list_multiple(splitinfo, splitsentence)
                                    for slidx, slval in enumerate(sublistindex):
                                        newsublistindex = list(range(slval[0], slval[1] + 1))
                                        if newsublistindex not in list(sublistindexdict.values()):
                                            newcombinedkey = combinedkey + '_' + str(slidx)
                                            sublistindexdict.update({newcombinedkey: newsublistindex})
                                            break

                    for infopoint in sublistindexdict:
                        try:
                            beginidx = sublistindexdict[infopoint][0]
                            endidx = sublistindexdict[infopoint][-1]
                        except IndexError:
                            print(sublistindexdict[infopoint])
                            print(infopoint)
                            print(sublistindexdict)
                            print(splitsentence)
                            print(sentence)
                            exit(0)
                        if beginidx != endidx:
                            nersentencepart = nersentence[beginidx:endidx+1]
                        else:
                            nersentencepart = [nersentence[beginidx]]
                        if all(x == 'O' for x in nersentencepart):
                            for actualidx, infopointidx in enumerate(sublistindexdict[infopoint]):
                                newkey = re.sub(r'\d+', '', infopoint)
                                newkey = re.sub(r'_', '', newkey)
                                if actualidx == 0:
                                    nertag = 'B-' + newkey.upper()
                                else:
                                    nertag = 'I-' + newkey.upper()
                                if nertag not in labels:
                                    labels.append(nertag)
                                nersentence[infopointidx] = nertag
                    if annotationtype == 'Prodigy':
                        allsentencelist.append([x[0] for x in splitsentence])
                    elif annotationtype == 'MyTool':
                        allsentencelist.append(splitsentence)
                    allnersentencelist.append(nersentence)

        finalstring = ''
        for idx, sentence in enumerate(allsentencelist):
            for idx2, word in enumerate(sentence):
                finalstring += allsentencelist[idx][idx2] + ' ' + allnersentencelist[idx][idx2] + '\n'
            finalstring += '\n'

        currentpath = os.getcwd()
        with open(currentpath + '/Data/NERFormat' + datasetlist[datasetidx] + '.txt', 'wb') as f:
            f.write(bytes(finalstring, 'UTF-8'))

    newlabels = '\n'.join(labels)
    currentpath = os.getcwd()
    with open(currentpath + '/Data/NERFormatLabels.txt', 'wb') as f:
        f.write(bytes(newlabels, 'UTF-8'))

def trecqa(alldatasets, relevantinfo):
    #First get a newrelevantinfo dict without the numbered infopoints (e.g. victim2Age, victim4Based, etc.)
    newrelevantinfo = {}

    for infopoint in relevantinfo:
        if (infopoint != 'sentenceidx') and (infopoint != 'filename') and (infopoint != 'sentence') and (infopoint != 'title') and (infopoint != 'textidx') and (infopoint != 'paragraphidx') and (infopoint != 'datasetidx'):
            newinfopoint = re.sub(r'\d+', '', infopoint)
            if newinfopoint not in newrelevantinfo:
                newrelevantinfo.update({newinfopoint: relevantinfo[infopoint]})

    datasetlist = ['Train', 'Dev', 'Test']
    #Go over the datasets
    for datasetidx, dataset in enumerate(alldatasets):
        trecdatasetlist = []
        #And the texts within the datasets
        for text in dataset:
            trectextlist = []
            #Go over each of the questions for each text
            for question in newrelevantinfo:
                for paragraph in text:
                    #Check for each sentence if the question we're looking at is answered
                    for sentence in paragraph:
                        if 'paragraphidx' in sentence:
                            annotationtype = 'Prodigy'
                        else:
                            annotationtype = 'MyTool'

                        combinedkeylist = []
                        #Remove digits from the keys and store them in combinedkeylist
                        for combinedkey in sentence:
                            newkey = re.sub(r'\d+', '', combinedkey)
                            if newkey not in combinedkeylist:
                                combinedkeylist.append(newkey)
                        #If the current question is answered, store with a 1 label, else with a 0 label
                        if question in combinedkeylist:
                            if annotationtype == 'MyTool':
                                trectextlist.append([question, 1, sentence['sentence'], sentence['sentenceidx'][0], sentence['sentenceidx'][1], sentence['sentenceidx'][2], sentence['sentenceidx'][3]])
                            elif annotationtype == 'Prodigy':
                                trectextlist.append([question, 1, sentence['sentence'], sentence['datasetidx'], sentence['textidx'], sentence['paragraphidx'], sentence['sentenceidx']])
                        else:
                            if annotationtype == 'MyTool':
                                trectextlist.append([question, 0, sentence['sentence'], sentence['sentenceidx'][0], sentence['sentenceidx'][1], sentence['sentenceidx'][2], sentence['sentenceidx'][3]])
                            elif annotationtype == 'Prodigy':
                                trectextlist.append([question, 0, sentence['sentence'], sentence['datasetidx'], sentence['textidx'], sentence['paragraphidx'], sentence['sentenceidx']])
            trecdatasetlist = trecdatasetlist + trectextlist

        wb = openpyxl.Workbook()
        ws = wb['Sheet']
        ws.append(['qtext', 'label', 'atext', 'dataset', 'text', 'paragraph', 'sentence'])
        for sentence in trecdatasetlist:
            ws.append(sentence)
        currentpath = os.getcwd()

        wb.save(currentpath + '/Data/TRECQAFormat' + datasetlist[datasetidx] + '.xlsx')

def classifier(alldatasets):
    currentpath = os.getcwd()
    classifierkeylist = []

    for dataset in alldatasets:
        for text in dataset:
            for paragraph in text:
                for sentence in paragraph:
                    for infopoint in sentence:
                        if (infopoint != 'sentenceidx') and (infopoint != 'filename') and (infopoint != 'sentence') and (infopoint != 'title') and (infopoint != 'textidx') and (infopoint != 'paragraphidx') and (infopoint != 'datasetidx'):
                            if infopoint not in classifierkeylist:
                                classifierkeylist.append(infopoint)

    classifiertextdict = {}
    for classkey in classifierkeylist:
        classifiertextdict.update({classkey: []})

    datasetlist = ['Train', 'Dev', 'Test']

    for datasetidx, dataset in enumerate(alldatasets):
        newclassdict = copy.deepcopy(classifiertextdict)
        for text in dataset:
            for paragraph in text:
                for sentence in paragraph:
                    if 'paragraphidx' in sentence:
                        annotationtype = 'Prodigy'
                    else:
                        annotationtype = 'MyTool'
                    combinedkeylist = []
                    for combinedkey in sentence:
                        combinedkeylist.append(combinedkey)
                    for question in newclassdict:
                        if question in combinedkeylist:
                            if annotationtype == 'MyTool':
                                newclassdict[question].append([1, sentence['sentence'], sentence['sentenceidx'][0], sentence['sentenceidx'][1], sentence['sentenceidx'][2], sentence['sentenceidx'][3]])
                            elif annotationtype == 'Prodigy':
                                newclassdict[question].append([1, sentence['sentence'], sentence['datasetidx'], sentence['textidx'], sentence['paragraphidx'], sentence['sentenceidx']])
                        else:
                            if annotationtype == 'MyTool':
                                newclassdict[question].append([0, sentence['sentence'], sentence['sentenceidx'][0], sentence['sentenceidx'][1], sentence['sentenceidx'][2], sentence['sentenceidx'][3]])
                            elif annotationtype == 'Prodigy':
                                newclassdict[question].append([0, sentence['sentence'], sentence['datasetidx'], sentence['textidx'], sentence['paragraphidx'], sentence['sentenceidx']])

        for questiontypeidx, questiontype in enumerate(newclassdict):
            if questiontypeidx == 0:
                wb = openpyxl.Workbook()
                del wb['Sheet']
            #else:
                #wb = openpyxl.load_workbook(currentpath + '/Data/ClassifierFormat' + datasetlist[datasetidx] + '.xlsx')

            wb.create_sheet(questiontype)
            ws = wb[questiontype]
            ws.append(['label', 'sentence', 'dataset', 'text', 'paragraph', 'sentence'])
            for sentence in newclassdict[questiontype]:
                ws.append(sentence)

            if questiontypeidx == len(newclassdict) -1:
                wb.save(currentpath + '/Data/ClassifierFormat' + datasetlist[datasetidx] + '.xlsx')

squaddataset(alldatasets)
webnlg(alldatasets, 'en')
#ner(alldatasets, 'en')
trecqa(alldatasets, relevantinfo)
classifier(alldatasets)