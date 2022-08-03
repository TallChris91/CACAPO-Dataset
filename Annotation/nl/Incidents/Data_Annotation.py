import tkinter as tk
from tkinter import messagebox
import os
import zipfile
import regex as re
import pickle
import json
import pyperclip
import openpyxl
from functools import partial

currentpath = os.getcwd()

with open(currentpath + '/Data/Textlist.pkl', 'rb') as f:
    alltexts = pickle.load(f)

with open(currentpath + '/Data/Filenamelist.pkl', 'rb') as f:
    allfiles = pickle.load(f)

def next_text():
    #Use a pickle list with format [datasetidx, articleidx, sentenceidx] to go to the last processed sentence
    with open(currentpath + '/Data/Donelist.pkl', 'rb') as f:
        textindices = pickle.load(f)

    if textindices == []:
        newindices = [0, 0, 0, 0]
        return newindices

    else:
        textindices = textindices[-1]

        newindices = [textindices[0], textindices[1], textindices[2], textindices[3] + 1]

        # If the last processed sentence was the last sentence of the paragraph
        if newindices[3] == len(alltexts[textindices[0]][textindices[1]][textindices[2]]):
            # Then go to the first sentence of the next paragraph
            newindices = [newindices[0], newindices[1], newindices[2] + 1, 0]
            # If the last processed sentence was the last sentence of the last text, go to the next text
            if newindices[2] == len(alltexts[textindices[0]][textindices[1]]):
                newindices = [newindices[0], newindices[1] + 1, 0, 0]
                # If the last processed sentence was the last sentence of the last text of the train/test/dev set, try to go to the next set
                if newindices[1] == len(alltexts[newindices[0]]):
                    newindices = [newindices[0] + 1, 0, 0, 0]
                    # If the last processed sentence was the overall last sentence of the corpus, we can shut this party down
                    if newindices[0] == len(newindices):
                        print("HOORAY, WE'RE DONE!")
                        exit(0)

    return newindices

def next_text_new(textindices):
    # Use a pickle list with format [datasetidx, articleidx, sentenceidx] to go to the last processed sentence
    if not os.path.exists(currentpath + '/Data/Donelist.pkl'):
        return 'New'

    newindices = [textindices[0], textindices[1], textindices[2], textindices[3] + 1]

    # If the last processed sentence was the last sentence of the paragraph
    if newindices[3] == len(alltexts[textindices[0]][textindices[1]][textindices[2]]):
        # Then go to the first sentence of the next paragraph
        newindices = [newindices[0], newindices[1], newindices[2] + 1, 0]
        # If the last processed sentence was the last sentence of the last text, go to the next text
        if newindices[2] == len(alltexts[textindices[0]][textindices[1]]):
            newindices = [newindices[0], newindices[1] + 1, 0, 0]
            # If the last processed sentence was the last sentence of the last text of the train/test/dev set, try to go to the next set
            if newindices[1] == len(alltexts[newindices[0]]):
                newindices = [newindices[0] + 1, 0, 0, 0]
                # If the last processed sentence was the overall last sentence of the corpus, we can shut this party down
                return 'Final textset'
            else:
                return 'Final paragraph'
        else:
            return 'Final sentence'
    else:
        return 'Not final'

def skip_text_new(textindices):
    newindices = [textindices[0], textindices[1] + 1, 0, 0]

    # If the last processed sentence was the last sentence of the last text of the train/test/dev set, try to go to the next set
    if newindices[1] == len(alltexts[newindices[0]]):
        newindices = [newindices[0] + 1, 0, 0, 0]
        # If the last processed sentence was the overall last sentence of the corpus, we can shut this party down
        return 'Final textset'

def choose_text(newindices, psentences='y'):
    if psentences == 'y':
        if os.path.exists(currentpath + '/Data/Tempprevioussentences.pkl'):
            with open(currentpath + '/Data/Tempprevioussentences.pkl', 'rb') as f:
                previoussentences = pickle.load(f)
            if len(previoussentences) > 0:
                previoussentences = ' '.join(previoussentences) + ' '
            else:
                previoussentences = ''
        else:
            previoussentences = ''
    try:
        currentsentence = alltexts[newindices[0]][newindices[1]][newindices[2]][newindices[3]]
    except IndexError:
        print("Hooray, we're done!")
        exit(0)

    #First add all sentences of the current paragraph
    if newindices[3] != len(alltexts[newindices[0]][newindices[1]][newindices[2]])-1:
        nextsentences = ' '.join(alltexts[newindices[0]][newindices[1]][newindices[2]][newindices[3]+1:])
    else:
        nextsentences = ''
    #Then add all sentences of the upcoming paragraphs
    if newindices[2] != len(alltexts[newindices[0]][newindices[1]]) - 1:
        for paragraph in alltexts[newindices[0]][newindices[1]][newindices[2]+1:]:
            paragraphsentences = ' '.join(paragraph)
            nextsentences = nextsentences + '\n\n' + paragraphsentences

    if psentences == 'y':
        return previoussentences, currentsentence, nextsentences
    else:
        return currentsentence, nextsentences

def get_jsonfile(newindices):
    currentfoldername = re.search(r'(?!.*\\)(.*?)$', currentpath).group(1)
    zipfilenames = ['Train.zip', 'Dev.zip', 'Test.zip']
    filename = allfiles[newindices[0]][newindices[1]]

    return filename

def startinfo(newindices, psentences='y'):
    if psentences == 'n':
        currentsentence, nextsentences = choose_text(newindices, 'n')
        filename = get_jsonfile(newindices)

        return currentsentence, nextsentences, filename
    else:
        previoussentences, currentsentence, nextsentences = choose_text(newindices)
        filename = get_jsonfile(newindices)

        return previoussentences, currentsentence, nextsentences, filename

def copy_new_sentence():
    currentsentence_field.delete(1.0, tk.END)
    currentsentence_field.insert(tk.END, pyperclip.paste())

def clear_add_datatype_field(event):
   add_datatype_field.delete(0, tk.END)

def clear_add_datatype_sentence_field(event):
   add_datatype_sentence_field.delete(0, tk.END)

def clear_add_datatype_entry_sentence_field(event):
   add_datatype_entry_sentence_field.delete(0, tk.END)

def clear_remove_datatype_field(event):
   remove_datatype_field.delete(0, tk.END)

def remove_datatype():
    global newrelevantinfo
    # remove previous IntVars
    entry_dict.clear()
    optionsmenu_dict.clear()

    # remove previous Checkboxes
    for rl in removelist:
        rl.destroy()
    removelist.clear()

    try:
        optionsmenu_dict['start'] = tk.StringVar()
        datatype_select_menu = tk.OptionMenu(gui, optionsmenu_dict['start'], *list(relevantinfo.keys()), command=add_box1)
        datatype_select_menu.grid(row=7, column=0, sticky='nesw')
        removelist.append(datatype_select_menu)
    except TypeError:
        ''

    with open(currentpath + '/Data/Relevantinfo.pkl', 'rb') as f:
        newrelevantinfo = pickle.load(f)

def remove_last(value):
    global newrelevantinfo

    lastkey = list(optionsmenu_dict.keys())[-1]
    del optionsmenu_dict[lastkey]

    removelist[-1].destroy()
    del removelist[-1]

    newrelevantinfo = relevantinfo.copy()
    for key in optionsmenu_dict:
        val = optionsmenu_dict[key].get()
        del newrelevantinfo[val]

    try:
        optionsmenu_dict[value] = tk.StringVar()
        datatype_select_menu = tk.OptionMenu(gui, optionsmenu_dict[value], *list(newrelevantinfo.keys()), command=add_box1)
        datatype_select_menu.grid(row=7 + len(entry_dict), column=0, sticky='nesw')
        removelist.append(datatype_select_menu)
    except TypeError:
        ''

    with open(currentpath + '/Data/Relevantinfo.pkl', 'rb') as f:
        newrelevantinfo = pickle.load(f)

def add_box1(value):
    global newrelevantinfo, entry_dict, optionsmenu_dict
    try:
        del newrelevantinfo[value]
    except ValueError:
        ''

    try:
        lastoptionsmenu = list(optionsmenu_dict.keys())[-1]
        lastoptionsmenuval = optionsmenu_dict[lastoptionsmenu].get()
    except IndexError:
        lastoptionsmenuval = None

    if lastoptionsmenuval == '':
        remove_last(value)
    else:
        entry_dict[value] = tk.StringVar()

        datatype_entry_1_field = tk.Entry(gui, textvariable=entry_dict[value])
        datatype_entry_1_field.grid(row=6 + len(entry_dict), column=1, columnspan=4, sticky='nesw')
        removelist.append(datatype_entry_1_field)

        try:
            optionsmenu_dict[value] = tk.StringVar()
            datatype_select_menu = tk.OptionMenu(gui, optionsmenu_dict[value], *list(newrelevantinfo.keys()), command=add_box1)
            datatype_select_menu.grid(row=7 + len(entry_dict), column=0, sticky='nesw')
            removelist.append(datatype_select_menu)
        except TypeError:
            ''


def add_new_datatype():
    global relevantinfo, newrelevantinfo, entrydict, optionsmenu_dict, removelist
    if (' ' not in add_datatype_field.get()) and (add_datatype_field.get() not in relevantinfo) and (add_datatype_field.get() != ''):
        newdatatype = add_datatype_field.get()
        relevantinfo.update({newdatatype: newdatatype})
        relevantinfo = dict(sorted(relevantinfo.items()))
        with open(currentpath + '/Data/Relevantinfo.pkl', 'wb') as f:
            pickle.dump(relevantinfo, f)

        try:
            # Reset var and delete all old options
            optionsmenu_dict['start'].set('')
            datatype_select_menu['menu'].delete(0, 'end')

            # Insert list of new options (tk._setit hooks them up to var)
            for choice in relevantinfo:
                datatype_select_menu['menu'].add_command(label=choice, command=tk._setit(optionsmenu_dict[list(optionsmenu_dict.keys())[0]], choice))
        except UnboundLocalError:
            optionsmenu_dict['start'] = tk.StringVar()
            datatype_select_menu = tk.OptionMenu(gui, optionsmenu_dict['start'], *list(relevantinfo.keys()), command=add_box1)
            datatype_select_menu.grid(row=7, column=0, sticky='nesw')
            removelist.append(datatype_select_menu)

        add_datatype_field.delete(0, tk.END)
        add_datatype_field.insert(tk.END, newdatatype + ' toegevoegd')

        # remove previous IntVars
        entry_dict.clear()
        optionsmenu_dict.clear()

        # remove previous Checkboxes
        for rl in removelist:
            rl.destroy()
        removelist.clear()

        try:
            optionsmenu_dict['start'] = tk.StringVar()
            datatype_select_menu = tk.OptionMenu(gui, optionsmenu_dict['start'], *list(relevantinfo.keys()), command=add_box1)
            datatype_select_menu.grid(row=7, column=0, sticky='nesw')
            removelist.append(datatype_select_menu)
        except TypeError:
            ''

        with open(currentpath + '/Data/Relevantinfo.pkl', 'rb') as f:
            newrelevantinfo = pickle.load(f)

    elif ' ' in add_datatype_field.get():
        add_datatype_field.delete(0, tk.END)
        add_datatype_field.insert(tk.END, 'Kan niet toevoegen als er spaties in staan')
    else:
        newdatatype = add_datatype_field.get()
        add_datatype_field.delete(0, tk.END)
        add_datatype_field.insert(tk.END, newdatatype + ' al in de lijst')

def remove_datatype_selection():
    global relevantinfo, newrelevantinfo, entrydict, optionsmenu_dict, removelist
    if (' ' not in remove_datatype_field.get()) and (remove_datatype_field.get() in relevantinfo):
        newdatatype = remove_datatype_field.get()
        del relevantinfo[newdatatype]
        relevantinfo = dict(sorted(relevantinfo.items()))
        with open(currentpath + '/Data/Relevantinfo.pkl', 'wb') as f:
            pickle.dump(relevantinfo, f)

        try:
            # Reset var and delete all old options
            optionsmenu_dict['start'].set('')
            datatype_select_menu['menu'].delete(0, 'end')

            # Insert list of new options (tk._setit hooks them up to var)
            for choice in relevantinfo:
                datatype_select_menu['menu'].add_command(label=choice, command=tk._setit(optionsmenu_dict[list(optionsmenu_dict.keys())[0]], choice))
        except UnboundLocalError:
            optionsmenu_dict['start'] = tk.StringVar()
            datatype_select_menu = tk.OptionMenu(gui, optionsmenu_dict['start'], *list(relevantinfo.keys()), command=add_box1)
            datatype_select_menu.grid(row=7, column=0, sticky='nesw')
            removelist.append(datatype_select_menu)

        remove_datatype_field.delete(0, tk.END)
        remove_datatype_field.insert(tk.END, newdatatype + ' weggehaald')

        # remove previous IntVars
        entry_dict.clear()
        optionsmenu_dict.clear()

        # remove previous Checkboxes
        for rl in removelist:
            rl.destroy()
        removelist.clear()

        try:
            optionsmenu_dict['start'] = tk.StringVar()
            datatype_select_menu = tk.OptionMenu(gui, optionsmenu_dict['start'], *list(relevantinfo.keys()), command=add_box1)
            datatype_select_menu.grid(row=7, column=0, sticky='nesw')
            removelist.append(datatype_select_menu)
        except TypeError:
            ''

        with open(currentpath + '/Data/Relevantinfo.pkl', 'rb') as f:
            newrelevantinfo = pickle.load(f)

    elif ' ' in remove_datatype_field.get():
        remove_datatype_field.delete(0, tk.END)
        remove_datatype_field.insert(tk.END, 'Kan niet toevoegen als er spaties in staan')
    else:
        newdatatype = remove_datatype_field.get()
        remove_datatype_field.delete(0, tk.END)
        remove_datatype_field.insert(tk.END, newdatatype + ' staat niet in de lijst')

def previous_sentence():
    global allindices, relevantinfo, newrelevantinfo, entrydict, optionsmenu_dict, removelist
    # Use a list of lists with format [datasetidx, articleidx, paragraphidx, sentenceidx] denoting all the processed sentences
    #And we can also only go to the previous datapoint if the donelist has done something
    if len(allindices) > 0:
        #Delete the current sentence
        del allindices[-1]

        #And go to the previous sentence
        textindices = allindices[-1]

        # And now delete this previous sentence from the allindices list and save this new index situation
        del allindices[-1]

        with open(currentpath + '/Data/Donelist.pkl', 'wb') as f:
            pickle.dump(allindices, f)

        #And add it again so that saving works as it should
        allindices.append(textindices)

        try:
            #Also open the Excelfile with the saved sentences and delete the last row from the excel list
            wb = openpyxl.load_workbook(currentpath + '/Data/Sentenceinfo.xlsx')
            ws = wb.active
            maxrow = ws.max_row
            ws.delete_rows(maxrow)
            wb.save(currentpath + '/Data/Sentenceinfo.xlsx')
        except FileNotFoundError:
            ''

        #Delete the last saved sentence from the tempparagraph as well
        with open(currentpath + '/Data/Tempparagraph.pkl', 'rb') as f:
            sentencelist = pickle.load(f)

        #If there are no sentences in the tempparagraph, the previous sentence is the last sentence of the previous paragraph
        if len(sentencelist) == 0:
            with open(currentpath + '/Data/Temptext.pkl', 'rb') as f:
                textlist = pickle.load(f)

            #If there are no sentences in the temptext, the previous sentence is the last sentence of the previous text
            if len(textlist) == 0:
                with open(currentpath + '/Data/Tempdataset.pkl', 'rb') as f:
                    datasetlist = pickle.load(f)

                # If there are no sentences in the tempdataset, the previous sentence is the last sentence of the last text of the train/dev/test set
                if len(datasetlist) == 0:
                    with open(currentpath + '/Data/Alldatasets.pkl', 'rb') as f:
                        alldatasetslist = pickle.load(f)

                    #Get the previous datasetlist
                    datasetlist = alldatasetslist[-1]
                    #And delete the previous datasetlist
                    del alldatasetslist[-1]
                    with open(currentpath + '/Data/Alldatasets.pkl', 'rb') as f:
                        pickle.dump(alldatasetslist, f)

                    # Get the last text
                    textlist = datasetlist[-1]
                    # Delete the last text from the tempdataset pickle
                    del datasetlist[-1]
                    with open(currentpath + '/Data/Tempdataset.pkl', 'wb') as f:
                        pickle.dump(datasetlist, f)

                    # Go to the last paragraph of the text
                    sentencelist = textlist[-1]
                    # And delete that last paragraph from the textlist
                    del textlist[-1]
                    with open(currentpath + '/Data/Temptext.pkl', 'wb') as f:
                        pickle.dump(textlist, f)

                    # And delete the last sentence from the last paragraph, as it's the current sentence now
                    del sentencelist[-1]
                    with open(currentpath + '/Data/Tempparagraph.pkl', 'wb') as f:
                        pickle.dump(sentencelist, f)

                    # Get the previoussentences from the information in the temptext
                    previoussentences = []

                    # Start with the previous paragraphs
                    for paragraph in textlist:
                        for combineddict in paragraph:
                            previoussentences.append(combineddict['sentence'])

                    # And then the sentences of the current paragraph
                    for sentence in sentencelist:
                        previoussentences.append(sentence['sentence'])

                    with open(currentpath + '/Data/Tempprevioussentences.pkl', 'wb') as f:
                        pickle.dump(previoussentences, f)

                #The previous sentence was the last sentence of the last text, so we'll go back to that situation
                else:
                    #Get the last text
                    textlist = datasetlist[-1]
                    #Delete the last text from the tempdataset pickle
                    del datasetlist[-1]
                    with open(currentpath + '/Data/Tempdataset.pkl', 'wb') as f:
                        pickle.dump(datasetlist, f)

                    #Go to the last paragraph of the text
                    sentencelist = textlist[-1]
                    #And delete that last paragraph from the textlist
                    del textlist[-1]
                    with open(currentpath + '/Data/Temptext.pkl', 'wb') as f:
                        pickle.dump(textlist, f)

                    #And delete the last sentence from the last paragraph, as it's the current sentence now
                    del sentencelist[-1]
                    with open(currentpath + '/Data/Tempparagraph.pkl', 'wb') as f:
                        pickle.dump(sentencelist, f)

                    # Get the previoussentences from the information in the temptext
                    previoussentences = []

                    # Start with the previous paragraphs
                    for paragraph in textlist:
                        for combineddict in paragraph:
                            previoussentences.append(combineddict['sentence'])

                    # And then the sentences of the current paragraph
                    for sentence in sentencelist:
                        previoussentences.append(sentence['sentence'])

                    with open(currentpath + '/Data/Tempprevioussentences.pkl', 'wb') as f:
                        pickle.dump(previoussentences, f)

            #Otherwise it is not the last sentence of the text and we can make the last processed paragraph the current paragraph
            else:
                sentencelist = textlist[-1]

                #Delete the last processed paragraph, as it is now the active paragraph again
                del textlist[-1]
                with open(currentpath + '/Data/Temptext.pkl', 'wb') as f:
                    pickle.dump(textlist, f)

                #Delete the last processed sentence of the paragraph
                del sentencelist[-1]
                with open(currentpath + '/Data/Tempparagraph.pkl', 'wb') as f:
                    pickle.dump(sentencelist, f)

                #Get the previoussentences from the information in the temptext
                previoussentences = []

                #Start with the previous paragraphs
                for paragraph in textlist:
                   for combineddict in paragraph:
                       previoussentences.append(combineddict['sentence'])

                #And then the sentences of the current paragraph
                for sentence in sentencelist:
                    previoussentences.append(sentence['sentence'])

                with open(currentpath + '/Data/Tempprevioussentences.pkl', 'wb') as f:
                    pickle.dump(previoussentences, f)

        #Otherwise, it's not the last sentence of the paragraph and we can delete the last processed sentence from the sentencelist
        else:
            del sentencelist[-1]
            with open(currentpath + '/Data/Tempparagraph.pkl', 'wb') as f:
                pickle.dump(sentencelist, f)

            #And delete the last sentence from the previoussentences
            with open(currentpath + '/Data/Tempprevioussentences.pkl', 'rb') as f:
                previoussentences = pickle.load(f)

            del previoussentences[-1]
            with open(currentpath + '/Data/Tempprevioussentences.pkl', 'wb') as f:
                pickle.dump(previoussentences, f)

        prevsents = ' '.join(previoussentences)

        currentsentence, nextsentences, filename = startinfo(textindices, 'n')

        results_field.delete(1.0, tk.END)
        results_field.insert(tk.END, prevsents)
        if (prevsents != '') and (nextsentences != ''):
            results_field.insert(tk.END, ' ' + currentsentence + ' ', "bold")
        elif (prevsents == '') and (nextsentences != ''):
            results_field.insert(tk.END, currentsentence + ' ', "bold")
        elif (prevsents != '') and (nextsentences == ''):
            results_field.insert(tk.END, ' ' + currentsentence, "bold")
        else:
            results_field.insert(tk.END, currentsentence, "bold")
        results_field.insert(tk.END, nextsentences)

        currentsentence_field.delete(1.0, tk.END)
        currentsentence_field.insert(tk.END, currentsentence)

        # remove previous entries and optionsdict entries
        entry_dict.clear()

        # remove previous optionsmenus
        for rl in removelist:
            rl.destroy()
        removelist.clear()

        optionsmenu_dict.clear()

        try:
            optionsmenu_dict['start'] = tk.StringVar()
            datatype_select_menu = tk.OptionMenu(gui, optionsmenu_dict['start'], *list(relevantinfo.keys()), command=add_box1)
            datatype_select_menu.grid(row=7, column=0, sticky='nesw')
            removelist.append(datatype_select_menu)
        except TypeError:
            ''

    with open(currentpath + '/Data/Relevantinfo.pkl', 'rb') as f:
        newrelevantinfo = pickle.load(f)

    zipfilenames = ['Train.zip', 'Dev.zip', 'Test.zip']
    if textindices[0] == 0:
        titlename = zipfilenames[textindices[0]] + ' ' + filename + ' (' + str(textindices[1] + 1) + '/100)'
    else:
        titlename = zipfilenames[textindices[0]] + ' ' + filename + ' (' + str(textindices[1] + 1) + '/50)'
    gui.title(titlename)

def skip_sentence():
    global allindices, previoussentences, currentsentence, nextsentences, relevantinfo, newrelevantinfo, filename, entrydict, optionsmenu_dict, removelist
    newstatus = next_text_new(allindices[-1])

    MsgBox = tk.messagebox.askquestion('Zin overslaan', 'Weet je zeker dat je de zin wil overslaan?',
                                       icon='warning')
    if MsgBox == 'yes':
        if newstatus == 'Final sentence':
            with open(currentpath + '/Data/Tempparagraph.pkl', 'rb') as f:
                sentencelist = pickle.load(f)

            with open(currentpath + '/Data/Temptext.pkl', 'rb') as f:
                textlist = pickle.load(f)

            textlist.append(sentencelist)

            with open(currentpath + '/Data/Temptext.pkl', 'wb') as f:
                pickle.dump(textlist, f)

            sentencelist = []
            with open(currentpath + '/Data/Tempparagraph.pkl', 'wb') as f:
                pickle.dump(sentencelist, f)

            with open(currentpath + '/Data/Tempprevioussentences.pkl', 'rb') as f:
                previoussentences = pickle.load(f)

            if len(previoussentences) > 0:
                lastsentence = previoussentences[-1] + '\n'
                del previoussentences[-1]
                previoussentences.append(lastsentence)

                with open(currentpath + '/Data/Tempprevioussentences.pkl', 'wb') as f:
                    pickle.dump(previoussentences, f)

        elif newstatus == 'Final paragraph':
            with open(currentpath + '/Data/Tempparagraph.pkl', 'rb') as f:
                sentencelist = pickle.load(f)

            with open(currentpath + '/Data/Temptext.pkl', 'rb') as f:
                textlist = pickle.load(f)

            textlist.append(sentencelist)

            with open(currentpath + '/Data/Tempdataset.pkl', 'rb') as f:
                datasetlist = pickle.load(f)

            datasetlist.append(textlist)

            with open(currentpath + '/Data/Tempdataset.pkl', 'wb') as f:
                pickle.dump(datasetlist, f)

            sentencelist = []
            with open(currentpath + '/Data/Tempparagraph.pkl', 'wb') as f:
                pickle.dump(sentencelist, f)

            textlist = []
            with open(currentpath + '/Data/Temptext.pkl', 'wb') as f:
                pickle.dump(textlist, f)

            with open(currentpath + '/Data/Temptitle.pkl', 'wb') as f:
                pickle.dump('', f)

            previoussentences = []

            with open(currentpath + '/Data/Tempprevioussentences.pkl', 'wb') as f:
                pickle.dump(previoussentences, f)

        elif newstatus == 'Final textset':
            with open(currentpath + '/Data/Tempparagraph.pkl', 'rb') as f:
                sentencelist = pickle.load(f)

            with open(currentpath + '/Data/Temptext.pkl', 'rb') as f:
                textlist = pickle.load(f)

            textlist.append(sentencelist)

            with open(currentpath + '/Data/Tempdataset.pkl', 'rb') as f:
                datasetlist = pickle.load(f)

            datasetlist.append(textlist)

            with open(currentpath + '/Data/Alldatasets.pkl', 'rb') as f:
                alldatasetslist = pickle.load(f)

            alldatasetslist.append(datasetlist)

            with open(currentpath + '/Data/Alldatasets.pkl', 'wb') as f:
                pickle.dump(alldatasetslist, f)

            sentencelist = []
            with open(currentpath + '/Data/Tempparagraph.pkl', 'wb') as f:
                pickle.dump(sentencelist, f)

            textlist = []
            with open(currentpath + '/Data/Temptext.pkl', 'wb') as f:
                pickle.dump(textlist, f)

            datasetlist = []
            with open(currentpath + '/Data/Tempdataset.pkl', 'wb') as f:
                pickle.dump(datasetlist, f)

            previoussentences = []

            with open(currentpath + '/Data/Tempprevioussentences.pkl', 'wb') as f:
                pickle.dump(previoussentences, f)

            with open(currentpath + '/Data/Temptitle.pkl', 'wb') as f:
                pickle.dump('', f)

        # remove previous entries and optionsdict entries
        entry_dict.clear()

        # remove previous optionsmenus
        for rl in removelist:
            rl.destroy()
        removelist.clear()

        optionsmenu_dict.clear()

        try:
            optionsmenu_dict['start'] = tk.StringVar()
            datatype_select_menu = tk.OptionMenu(gui, optionsmenu_dict['start'], *list(relevantinfo.keys()), command=add_box1)
            datatype_select_menu.grid(row=7, column=0, sticky='nesw')
            removelist.append(datatype_select_menu)
        except TypeError:
            ''

        save_field.delete(0, tk.END)
        if len(currentsentence_field.get("1.0",tk.END).strip()) >= 30:
            save_field.insert(tk.END, currentsentence_field.get("1.0",tk.END).strip()[:30] + '... overgeslagen')
        else:
            save_field.insert(tk.END, currentsentence_field.get("1.0", tk.END).strip() + ' overgeslagen')


        with open(currentpath + '/Data/Donelist.pkl', 'wb') as f:
            pickle.dump(allindices, f)

        newindices = next_text()
        allindices.append(newindices)
        previoussentences, currentsentence, nextsentences, filename = startinfo(allindices[-1])

        results_field.delete(1.0, tk.END)
        results_field.insert(tk.END, previoussentences)
        if (previoussentences != '') and (nextsentences != ''):
            results_field.insert(tk.END, ' ' + currentsentence + ' ', "bold")
        elif (previoussentences == '') and (nextsentences != ''):
            results_field.insert(tk.END, currentsentence + ' ', "bold")
        elif (previoussentences != '') and (nextsentences == ''):
            results_field.insert(tk.END, ' ' + currentsentence, "bold")
        else:
            results_field.insert(tk.END, currentsentence, "bold")
        results_field.insert(tk.END, nextsentences)

        currentsentence_field.delete(1.0, tk.END)
        currentsentence_field.insert(tk.END, currentsentence)

        add_datatype_field.delete(0, tk.END)
        add_datatype_field.insert(tk.END, 'Vul hier een nieuw datatype in')

        if not os.path.exists(currentpath + '/Data/Relevantinfo.pkl'):
            relevantinfo = {}
            with open(currentpath + '/Data/Relevantinfo.pkl', 'wb') as f:
                pickle.dump(relevantinfo, f)
        else:
            with open(currentpath + '/Data/Relevantinfo.pkl', 'rb') as f:
                relevantinfo = pickle.load(f)

        newrelevantinfo = relevantinfo.copy()

        zipfilenames = ['Train.zip', 'Dev.zip', 'Test.zip']
        if allindices[-1][0] == 0:
            titlename = zipfilenames[allindices[-1][0]] + ' ' + filename + ' (' + str(allindices[-1][1] + 1) + '/100)'
        else:
            titlename = zipfilenames[allindices[-1][0]] + ' ' + filename + ' (' + str(allindices[-1][1] + 1) + '/50)'
        gui.title(titlename)

def skip_text():
    global allindices, previoussentences, currentsentence, nextsentences, relevantinfo, newrelevantinfo, filename, entrydict, optionsmenu_dict, removelist
    newstatus = skip_text_new(allindices[-1])

    MsgBox = tk.messagebox.askquestion('Tekst overslaan', 'Weet je zeker dat je de tekst wil overslaan?',
                                       icon='warning')
    if MsgBox == 'yes':
        #If we figured out we should skip this text while having already annotated some data, we should clear this text information
        sentencelist = []
        with open(currentpath + '/Data/Tempparagraph.pkl', 'wb') as f:
            pickle.dump(sentencelist, f)

        textlist = []
        with open(currentpath + '/Data/Temptext.pkl', 'wb') as f:
            pickle.dump(textlist, f)

        previoussentences = []

        with open(currentpath + '/Data/Tempprevioussentences.pkl', 'wb') as f:
            pickle.dump(previoussentences, f)

        with open(currentpath + '/Data/Temptitle.pkl', 'wb') as f:
            pickle.dump('', f)

        #If this text is the final text of the dataset, we should still append the dataset containing all previous texts to the alldataset file
        if newstatus == 'Final textset':
            with open(currentpath + '/Data/Tempdataset.pkl', 'rb') as f:
                datasetlist = pickle.load(f)

            with open(currentpath + '/Data/Alldatasets.pkl', 'rb') as f:
                alldatasetslist = pickle.load(f)

            alldatasetslist.append(datasetlist)

            with open(currentpath + '/Data/Alldatasets.pkl', 'wb') as f:
                pickle.dump(alldatasetslist, f)

            datasetlist = []
            with open(currentpath + '/Data/Tempdataset.pkl', 'wb') as f:
                pickle.dump(datasetlist, f)

        # remove previous entries and optionsdict entries
        entry_dict.clear()

        # remove previous optionsmenus
        for rl in removelist:
            rl.destroy()
        removelist.clear()

        optionsmenu_dict.clear()

        try:
            optionsmenu_dict['start'] = tk.StringVar()
            datatype_select_menu = tk.OptionMenu(gui, optionsmenu_dict['start'], *list(relevantinfo.keys()), command=add_box1)
            datatype_select_menu.grid(row=7, column=0, sticky='nesw')
            removelist.append(datatype_select_menu)
        except TypeError:
            ''

        save_field.delete(0, tk.END)
        if len(currentsentence_field.get("1.0",tk.END).strip()) >= 30:
            save_field.insert(tk.END, filename + '... overgeslagen')
        else:
            save_field.insert(tk.END, filename + ' overgeslagen')

        #Get the number of the current text
        currenttextnum = allindices[-1][1]

        #Go to the new sentence until the next text is reached
        while allindices[-1][1] == currenttextnum:
            with open(currentpath + '/Data/Donelist.pkl', 'wb') as f:
                pickle.dump(allindices, f)

            newindices = next_text()
            allindices.append(newindices)

        previoussentences, currentsentence, nextsentences, filename = startinfo(allindices[-1])

        results_field.delete(1.0, tk.END)
        results_field.insert(tk.END, previoussentences)
        if (previoussentences != '') and (nextsentences != ''):
            results_field.insert(tk.END, ' ' + currentsentence + ' ', "bold")
        elif (previoussentences == '') and (nextsentences != ''):
            results_field.insert(tk.END, currentsentence + ' ', "bold")
        elif (previoussentences != '') and (nextsentences == ''):
            results_field.insert(tk.END, ' ' + currentsentence, "bold")
        else:
            results_field.insert(tk.END, currentsentence, "bold")
        results_field.insert(tk.END, nextsentences)

        currentsentence_field.delete(1.0, tk.END)
        currentsentence_field.insert(tk.END, currentsentence)

        add_datatype_field.delete(0, tk.END)
        add_datatype_field.insert(tk.END, 'Vul hier een nieuw datatype in')

        if not os.path.exists(currentpath + '/Data/Relevantinfo.pkl'):
            relevantinfo = {}
            with open(currentpath + '/Data/Relevantinfo.pkl', 'wb') as f:
                pickle.dump(relevantinfo, f)
        else:
            with open(currentpath + '/Data/Relevantinfo.pkl', 'rb') as f:
                relevantinfo = pickle.load(f)

        newrelevantinfo = relevantinfo.copy()

        zipfilenames = ['Train.zip', 'Dev.zip', 'Test.zip']
        if allindices[-1][0] == 0:
            titlename = zipfilenames[allindices[-1][0]] + ' ' + filename + ' (' + str(allindices[-1][1] + 1) + '/100)'
        else:
            titlename = zipfilenames[allindices[-1][0]] + ' ' + filename + ' (' + str(allindices[-1][1] + 1) + '/50)'
        gui.title(titlename)

def save_title():
    with open(currentpath + '/Data/Temptitle.pkl', 'wb') as f:
        pickle.dump(currentsentence_field.get("1.0", tk.END).strip(), f)
    save_field.delete(0, tk.END)
    save_field.insert(tk.END, currentsentence_field.get("1.0", tk.END).strip() + ' als titel opgeslagen')
    #skip_sentence()


def save_all():
    global allindices, previoussentences, currentsentence, nextsentences, relevantinfo, newrelevantinfo, filename, entrydict, optionsmenu_dict, removelist
    zipfilenames = ['Train.zip', 'Dev.zip', 'Test.zip']
    zipfilename = allindices[-1][0]

    enteredentries = []
    chosenoptions = []

    for key in entry_dict:
        enteredentries.append(entry_dict[key].get())

    for key in optionsmenu_dict:
        chosenoptions.append(optionsmenu_dict[key].get())

    enteredentries = [x for x in enteredentries if x != '']
    chosenoptions = [x for x in chosenoptions if x != '']

    if len(enteredentries) != len(chosenoptions):
        save_field.delete(0, tk.END)
        save_field.insert(tk.END, 'Zorg dat alle entries ingevuld zijn bij de datapunten!')
        return ''

    combinedlist = []
    combineddict = {}

    for idx, val in enumerate(enteredentries):
        combineddict.update({chosenoptions[idx]: enteredentries[idx]})
        combinedlist.append([chosenoptions[idx], enteredentries[idx]])

    newcombinedlist = [' | '.join(x) for x in combinedlist]

    combineddict.update({'sentenceidx': allindices[-1]})
    combineddict.update({'filename': filename})

    with open(currentpath + '/Data/Temptitle.pkl', 'rb') as f:
        texttitle = pickle.load(f)
    combineddict.update({'title': texttitle})

    newstatus = next_text_new(allindices[-1])
    if newstatus == 'Final sentence':
        excellist = [zipfilenames[zipfilename], filename, allindices[-1][0], allindices[-1][1], allindices[-1][2], allindices[-1][3],
                     currentsentence_field.get("1.0", tk.END).strip() + '\n'] + newcombinedlist

        combineddict.update({'sentence': currentsentence_field.get("1.0", tk.END).strip() + '\n'})
    else:
        excellist = [zipfilenames[zipfilename], filename, allindices[-1][0], allindices[-1][1], allindices[-1][2], allindices[-1][3],
                     currentsentence_field.get("1.0", tk.END).strip()] + newcombinedlist
        combineddict.update({'sentence': currentsentence_field.get("1.0", tk.END).strip()})

    if newstatus == 'New':
        wb = openpyxl.Workbook()
        ws = wb['Sheet']
        ws.append(excellist)

        wb.save(currentpath + '/Data/Sentenceinfo.xlsx')

        sentencelist = []

        sentencelist.append(combineddict)

        with open(currentpath + '/Data/Tempparagraph.pkl', 'wb') as f:
            pickle.dump(sentencelist, f)

        with open(currentpath + '/Data/Tempprevioussentences.pkl', 'wb') as f:
            pickle.dump([currentsentence_field.get("1.0",tk.END).strip()], f)

    elif newstatus == 'Not final':
        try:
            wb = openpyxl.load_workbook(currentpath + '/Data/Sentenceinfo.xlsx')
            ws = wb.active
            ws.append(excellist)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            ws = wb['Sheet']
            ws.append(excellist)

        wb.save(currentpath + '/Data/Sentenceinfo.xlsx')

        with open(currentpath + '/Data/Tempparagraph.pkl', 'rb') as f:
            sentencelist = pickle.load(f)

        sentencelist.append(combineddict)

        with open(currentpath + '/Data/Tempparagraph.pkl', 'wb') as f:
            pickle.dump(sentencelist, f)

        with open(currentpath + '/Data/Tempprevioussentences.pkl', 'rb') as f:
            previoussentences = pickle.load(f)

        previoussentences.append(currentsentence_field.get("1.0", tk.END).strip())

        with open(currentpath + '/Data/Tempprevioussentences.pkl', 'wb') as f:
            pickle.dump(previoussentences, f)

    #Last sentence of the paragraph
    elif newstatus == 'Final sentence':
        wb = openpyxl.load_workbook(currentpath + '/Data/Sentenceinfo.xlsx')
        ws = wb.active
        ws.append(excellist)
        wb.save(currentpath + '/Data/Sentenceinfo.xlsx')

        with open(currentpath + '/Data/Tempparagraph.pkl', 'rb') as f:
            sentencelist = pickle.load(f)

        sentencelist.append(combineddict)

        with open(currentpath + '/Data/Temptext.pkl', 'rb') as f:
            textlist = pickle.load(f)

        textlist.append(sentencelist)

        with open(currentpath + '/Data/Temptext.pkl', 'wb') as f:
            pickle.dump(textlist, f)

        sentencelist = []
        with open(currentpath + '/Data/Tempparagraph.pkl', 'wb') as f:
            pickle.dump(sentencelist, f)

        with open(currentpath + '/Data/Tempprevioussentences.pkl', 'rb') as f:
            previoussentences = pickle.load(f)

        previoussentences.append(currentsentence_field.get("1.0", tk.END).strip() + '\n\n')

        with open(currentpath + '/Data/Tempprevioussentences.pkl', 'wb') as f:
            pickle.dump(previoussentences, f)

    #Last sentence of the text
    elif newstatus == 'Final paragraph':
        wb = openpyxl.load_workbook(currentpath + '/Data/Sentenceinfo.xlsx')
        ws = wb.active
        ws.append(excellist)
        wb.save(currentpath + '/Data/Sentenceinfo.xlsx')

        with open(currentpath + '/Data/Tempparagraph.pkl', 'rb') as f:
            sentencelist = pickle.load(f)

        sentencelist.append(combineddict)

        with open(currentpath + '/Data/Temptext.pkl', 'rb') as f:
            textlist = pickle.load(f)

        textlist.append(sentencelist)

        with open(currentpath + '/Data/Tempdataset.pkl', 'rb') as f:
            datasetlist = pickle.load(f)

        datasetlist.append(textlist)

        with open(currentpath + '/Data/Tempdataset.pkl', 'wb') as f:
            pickle.dump(datasetlist, f)

        sentencelist = []
        with open(currentpath + '/Data/Tempparagraph.pkl', 'wb') as f:
            pickle.dump(sentencelist, f)

        textlist = []
        with open(currentpath + '/Data/Temptext.pkl', 'wb') as f:
            pickle.dump(textlist, f)

        previoussentences = []

        with open(currentpath + '/Data/Tempprevioussentences.pkl', 'wb') as f:
            pickle.dump(previoussentences, f)

        with open(currentpath + '/Data/Temptitle.pkl', 'wb') as f:
            pickle.dump('', f)

    #Last sentence of the final text in the train/dev/test set
    elif newstatus == 'Final textset':
        wb = openpyxl.load_workbook(currentpath + '/Data/Sentenceinfo.xlsx')
        ws = wb.active
        ws.append(excellist)
        wb.save(currentpath + '/Data/Sentenceinfo.xlsx')

        with open(currentpath + '/Data/Tempparagraph.pkl', 'rb') as f:
            sentencelist = pickle.load(f)

        sentencelist.append(combineddict)

        with open(currentpath + '/Data/Temptext.pkl', 'rb') as f:
            textlist = pickle.load(f)

        textlist.append(sentencelist)

        with open(currentpath + '/Data/Tempdataset.pkl', 'rb') as f:
            datasetlist = pickle.load(f)

        datasetlist.append(textlist)

        with open(currentpath + '/Data/Alldatasets.pkl', 'rb') as f:
            alldatasetslist = pickle.load(f)

        alldatasetslist.append(datasetlist)

        with open(currentpath + '/Data/Alldatasets.pkl', 'wb') as f:
            pickle.dump(alldatasetslist, f)

        sentencelist = []
        with open(currentpath + '/Data/Tempparagraph.pkl', 'wb') as f:
            pickle.dump(sentencelist, f)

        textlist = []
        with open(currentpath + '/Data/Temptext.pkl', 'wb') as f:
            pickle.dump(textlist, f)

        datasetlist = []
        with open(currentpath + '/Data/Tempdataset.pkl', 'wb') as f:
            pickle.dump(datasetlist, f)

        previoussentences = []

        with open(currentpath + '/Data/Tempprevioussentences.pkl', 'wb') as f:
            pickle.dump(previoussentences, f)

        with open(currentpath + '/Data/Temptitle.pkl', 'wb') as f:
            pickle.dump('', f)

    # remove previous entries and optionsdict entries
    entry_dict.clear()

    # remove previous optionsmenus
    for rl in removelist:
        rl.destroy()
    removelist.clear()

    optionsmenu_dict.clear()

    try:
        optionsmenu_dict['start'] = tk.StringVar()
        datatype_select_menu = tk.OptionMenu(gui, optionsmenu_dict['start'], *list(relevantinfo.keys()), command=add_box1)
        datatype_select_menu.grid(row=7, column=0, sticky='nesw')
        removelist.append(datatype_select_menu)
    except TypeError:
        ''

    save_field.delete(0, tk.END)
    if len(currentsentence_field.get("1.0",tk.END).strip()) >= 30:
        save_field.insert(tk.END, currentsentence_field.get("1.0",tk.END).strip()[:30] + '... toegevoegd aan de database')
    else:
        save_field.insert(tk.END, currentsentence_field.get("1.0", tk.END).strip() + ' toegevoegd aan de database')

    with open(currentpath + '/Data/Donelist.pkl', 'wb') as f:
        pickle.dump(allindices, f)

    newindices = next_text()
    allindices.append(newindices)
    previoussentences, currentsentence, nextsentences, filename = startinfo(allindices[-1])

    results_field.delete(1.0, tk.END)
    results_field.insert(tk.END, previoussentences)
    if (previoussentences != '') and (nextsentences != ''):
        results_field.insert(tk.END, ' ' + currentsentence + ' ', "bold")
    elif (previoussentences == '') and (nextsentences != ''):
        results_field.insert(tk.END, currentsentence + ' ', "bold")
    elif (previoussentences != '') and (nextsentences == ''):
        results_field.insert(tk.END, ' ' + currentsentence, "bold")
    else:
        results_field.insert(tk.END, currentsentence, "bold")
    results_field.insert(tk.END, nextsentences)

    currentsentence_field.delete(1.0, tk.END)
    currentsentence_field.insert(tk.END, currentsentence)

    add_datatype_field.delete(0, tk.END)
    add_datatype_field.insert(tk.END, 'Vul hier een nieuw datatype in')

    if not os.path.exists(currentpath + '/Data/Relevantinfo.pkl'):
        relevantinfo = {}
        with open(currentpath + '/Data/Relevantinfo.pkl', 'wb') as f:
            pickle.dump(relevantinfo, f)
    else:
        with open(currentpath + '/Data/Relevantinfo.pkl', 'rb') as f:
            relevantinfo = pickle.load(f)

    newrelevantinfo = relevantinfo.copy()

    zipfilenames = ['Train.zip', 'Dev.zip', 'Test.zip']
    if allindices[-1][0] == 0:
        titlename = zipfilenames[allindices[-1][0]] + ' ' + filename + ' (' + str(allindices[-1][1] + 1) + '/100)'
    else:
        titlename = zipfilenames[allindices[-1][0]] + ' ' + filename + ' (' + str(allindices[-1][1] + 1) + '/50)'
    gui.title(titlename)

if not os.path.exists(currentpath + '/Data/Donelist.pkl'):
    allindices = []
    with open(currentpath + '/Data/Donelist.pkl', 'wb') as f:
        pickle.dump(allindices, f)

with open(currentpath + '/Data/Donelist.pkl', 'rb') as f:
    allindices = pickle.load(f)
newindices = next_text()
allindices.append(newindices)
previoussentences, currentsentence, nextsentences, filename = startinfo(allindices[-1])

entry_dict = {}
optionsmenu_dict = {}
removelist = []

zipfilenames = ['Train.zip', 'Dev.zip', 'Test.zip']

if allindices[-1][0] == 0:
    titlename = zipfilenames[allindices[-1][0]] + ' ' + filename + ' (' + str(allindices[-1][1] + 1) + '/100)'
else:
    titlename = zipfilenames[allindices[-1][0]] + ' ' + filename + ' (' + str(allindices[-1][1] + 1) + '/50)'

if not os.path.exists(currentpath + '/Data/Tempparagraph.pkl'):
    sentencelist = []
    with open(currentpath + '/Data/Tempparagraph.pkl', 'wb') as f:
        pickle.dump(sentencelist, f)

if not os.path.exists(currentpath + '/Data/Temptext.pkl'):
    textlist = []
    with open(currentpath + '/Data/Temptext.pkl', 'wb') as f:
        pickle.dump(textlist, f)

if not os.path.exists(currentpath + '/Data/Tempdataset.pkl'):
    datasetlist = []
    with open(currentpath + '/Data/Tempdataset.pkl', 'wb') as f:
        pickle.dump(datasetlist, f)

if not os.path.exists(currentpath + '/Data/Alldatasets.pkl'):
    alldatasetslist = []

    with open(currentpath + '/Data/Alldatasets.pkl', 'wb') as f:
        pickle.dump(alldatasetslist, f)

if not os.path.exists(currentpath + '/Data/Tempprevioussentences.pkl'):
    previoussentences = []

    with open(currentpath + '/Data/Tempprevioussentences.pkl', 'wb') as f:
        pickle.dump(previoussentences, f)

if not os.path.exists(currentpath + '/Data/Temptitle.pkl'):
    with open(currentpath + '/Data/Temptitle.pkl', 'wb') as f:
        pickle.dump('', f)

# create a GUI window
gui = tk.Tk()

# set the title of GUI window
gui.title(titlename)

currentsentence_text = tk.Label(gui, text="Volledig bericht:")
currentsentence_text.grid(row=0, column=0, sticky=tk.W)

# create the text box for showing the full text with the current sentence bold.
scroll = tk.Scrollbar(gui)
results_field = tk.Text(gui, height=20, width=100, wrap=tk.WORD, font=('TkDefaultFont', 9))
results_field.tag_configure("bold", font=('TkDefaultFont', 9,'bold'))
results_field.grid(row=1, column=0, columnspan=5)
scroll.grid(row=1, column=6, sticky=tk.E)
scroll.config(command=results_field.yview)
results_field.config(yscrollcommand=scroll.set)

results_field.delete(1.0, tk.END)
results_field.insert(tk.END, previoussentences)
if (previoussentences != '') and (nextsentences != ''):
    results_field.insert(tk.END, ' ' + currentsentence + ' ', "bold")
elif (previoussentences == '') and (nextsentences != ''):
    results_field.insert(tk.END, currentsentence + ' ', "bold")
elif (previoussentences != '') and (nextsentences == ''):
    results_field.insert(tk.END, ' ' + currentsentence, "bold")
else:
    results_field.insert(tk.END, currentsentence, "bold")
results_field.insert(tk.END, nextsentences)

currentsentence_text = tk.Label(gui, text="Huidige zin:")
currentsentence_text.grid(row=2, column=0, sticky=tk.W)

# create the text box for showing an example response.
currentsentence_field = tk.Text(gui, height=3, width=100, wrap=tk.WORD, font=('TkDefaultFont', 9))
currentsentence_field.grid(row=3, column=0, pady=(10, 0), columnspan=5)

currentsentence_field.delete(1.0, tk.END)
currentsentence_field.insert(tk.END, currentsentence)

#copy_new_button = tk.Button(gui, text='Plak nieuwe tekst', command=copy_new_sentence)
#copy_new_button.grid(row=4, column=0, sticky='nesw')

if not os.path.exists(currentpath + '/Data/Relevantinfo.pkl'):
    relevantinfo = {}
    with open(currentpath + '/Data/Relevantinfo.pkl', 'wb') as f:
        pickle.dump(relevantinfo, f)
else:
    with open(currentpath + '/Data/Relevantinfo.pkl', 'rb') as f:
        relevantinfo = pickle.load(f)

newrelevantinfo = relevantinfo.copy()

add_datatype_button = tk.Button(gui, text='Voeg datatype toe', command=add_new_datatype)
add_datatype_button.grid(row=5, column=0, sticky='nesw')

add_datatype_var = tk.StringVar()
add_datatype_field = tk.Entry(gui, textvariable=add_datatype_var)
add_datatype_field.grid(row=5, column=1, sticky='nesw')
add_datatype_field.delete(0, tk.END)
add_datatype_field.insert(tk.END, 'Vul hier een nieuw datatype in')
add_datatype_field.bind("<Button-1>", clear_add_datatype_field)

remove_datatype_button = tk.Button(gui, text='Verwijder datatype selectie', command=remove_datatype)
remove_datatype_button.grid(row=6, column=0, sticky='nesw')

try:
    optionsmenu_dict['start'] = tk.StringVar()
    datatype_select_menu = tk.OptionMenu(gui, optionsmenu_dict['start'], *list(relevantinfo.keys()), command=add_box1)
    datatype_select_menu.grid(row=7, column=0, sticky='nesw')
    removelist.append(datatype_select_menu)
except TypeError:
    ''

save_button = tk.Button(gui, text='Opslaan', command=save_all)
save_button.grid(row=30, column=0, sticky='nesw')

save_title_button = tk.Button(gui, text='Als titel opslaan', command=save_title)
save_title_button.grid(row=30, column=1, sticky='nesw')

previous_button = tk.Button(gui, text='Vorige', command=previous_sentence)
previous_button.grid(row=30, column=2, sticky='nesw')

save_var = tk.StringVar()
save_field = tk.Entry(gui, textvariable=save_var)
save_field.grid(row=31, column=0, columnspan=4, sticky='nesw')
save_field.delete(0, tk.END)
save_field.insert(tk.END, 'Hier verschijnt het resultaat')

skip_sentence_button = tk.Button(gui, text='Sla zin over', command=skip_sentence)
skip_sentence_button.grid(row=32, column=0, pady=(20,0), sticky='nesw')

skip_text_button = tk.Button(gui, text='Sla text over', command=skip_text)
skip_text_button.grid(row=32, column=1, pady=(20,0), sticky='nesw')

# start the GUI
gui.mainloop()